import os
import logging
from exchangelib import Credentials, Account, DELEGATE, FileAttachment
import openpyxl
from datetime import datetime
import pandas as pd
import easygui
from tqdm import tqdm

def detectar_tipo_anexo(nome_anexo):
    palavras_nf = ["nf", "nota fiscal", "invoice"]
    palavras_certificado = ["certificado", "certification", "material certification"]

    if any(palavra in nome_anexo.lower() for palavra in palavras_nf):
        return "Nota Fiscal"
    elif any(palavra in nome_anexo.lower() for palavra in palavras_certificado):
        return "Certificação de Material"
    else:
        return "Outro"

def extrair_metadados_anexo(attachment):
    return {"Nome": attachment.name, "Tamanho": attachment.size}

def solicitar_opcoes_filtro():
    msg = "Escolha a pasta de e-mails:"
    title = "Opções de Filtro"
    choices = ["Caixa de Entrada", "Rascunhos", "Itens Enviados"]
    selected_options = easygui.choicebox(msg, title, choices)

    return selected_options

def encontrar_nome_arquivo_disponivel(pasta, nome_base):
    contador = 1
    nome_arquivo = f"{nome_base}.xlsx"

    while os.path.exists(os.path.join(pasta, nome_arquivo)):
        contador += 1
        nome_arquivo = f"{nome_base}_version_{contador}.xlsx"

    return nome_arquivo

def extrair_metadados_anexo(attachment):
    return {"Nome": attachment.name, "Tamanho": attachment.size}

def extrair_emails(account, selected_folder, progress_var, log_data):
    folder_mapping = {"Caixa de Entrada": account.inbox, "Rascunhos": account.drafts, "Itens Enviados": account.sent}
    folder = folder_mapping.get(selected_folder)

    if folder:
        emails = folder.all()
        total_emails = len(emails)
        
        for index, item in enumerate(emails, start=1):
            sender = item.sender.email_address
            to_recipients = ", ".join(recipient.email_address for recipient in item.to_recipients)
            cc_recipients = ", ".join(recipient.email_address for recipient in item.cc_recipients)
            subject = item.subject
            body = item.text_body
            received_date = item.datetime_received

            if item.has_attachments:
                for i, attachment in enumerate(item.attachments, start=1):
                    attachment_name = attachment.name
                    attachment_size = attachment.size

                    tipo_anexo = detectar_tipo_anexo(attachment_name)
                    metadados_anexo = extrair_metadados_anexo(attachment)

                    # Adiciona os dados à planilha
                    main_sheet.append([sender, to_recipients, cc_recipients, subject,
                                    f"{i} - {attachment_name}", f"{i} - {attachment.content}", body, received_date,
                                    tipo_anexo, metadados_anexo])
            else:
                main_sheet.append([sender, to_recipients, cc_recipients, subject, "", "", body, received_date, "", ""])

            log_data["Subject"].append(subject)
            log_data["Status"].append("Completo")

            # Atualiza a barra de progresso
            progress_var.update(index)

        logging.info(f"Total de {total_emails} e-mails processados com sucesso.")
    else:
        logging.error("Pasta de e-mails não reconhecida.")

def extrair_emails_thread(account, selected_folder, progress_var, log_data):
    try:
        extrair_emails(account, selected_folder, progress_var, log_data)
    except Exception as e:
        logging.error(f"Erro durante a execução do script: {str(e)}")
        for subject in log_data["Subject"]:
            log_data["Status"].append(f"Falha - {str(e)}")

def criar_planilha_e_log(data_mail_folder, nome_arquivo, main_sheet, log_sheet):
    # Salvar o arquivo com o nome encontrado
    wb = openpyxl.Workbook()
    wb.save(os.path.join(data_mail_folder, nome_arquivo))

    logging.info("Script concluído.")

    # Mostrar opção para abrir a planilha baixada
    if easygui.ynbox("Deseja abrir a planilha baixada?", title="Concluído", choices=("Sim", "Não")):
        os.startfile(os.path.join(data_mail_folder, nome_arquivo))

def solicitar_credenciais():
    msg = "Digite suas credenciais:"
    title = "Credenciais"
    field_names = ["Endereço de E-mail", "Senha"]
    field_values = easygui.multpasswordbox(msg, title, field_names)

    return field_values

def main():
    # Obter credenciais usando a nova função
    credentials = solicitar_credenciais()

    if credentials is None:
        print("Credenciais não fornecidas. Certifique-se de inserir um endereço de e-mail e senha válidos.")
    else:
        email_address, password = credentials
        credentials = Credentials(email_address, password)
        account = Account(email_address, credentials=credentials, autodiscover=True, access_type=DELEGATE)

        # Solicitar a pasta de e-mails
        selected_folder = solicitar_opcoes_filtro()

        # Solicitar o caminho para salvar os arquivos
        data_mail_folder = easygui.diropenbox("Escolha a pasta para salvar os arquivos:")
        
        # Nome do arquivo
        data_atual = datetime.now().strftime("%d%m%Y")
        nome_base_arquivo = f"Compilados_emails_{data_atual}"
        
        # Encontrar um nome de arquivo disponível
        nome_arquivo = encontrar_nome_arquivo_disponivel(data_mail_folder, nome_base_arquivo)

        # Configurar a interface gráfica para mostrar o progresso
        progress_max = 100  # Pode ser ajustado conforme necessário
        progress_var = tqdm(total=progress_max, desc="Baixando e-mails", unit=" e-mails", position=0)

        wb = openpyxl.Workbook()
        main_sheet = wb.active
        main_sheet.title = "E-mails"

        log_sheet = wb.create_sheet("Log de Atividades")

        main_sheet.append(["DE (From)", "PARA (To)", "CC (CC)", "TÍTULO DO E-MAIL (Email Subject)", "ANEXO (Attachment)",
                        "TEXTO DO ANEXO", "TEXTO (Body Text)", "DATA DE RECEBIMENTO", "TIPO DE ANEXO", "METADADOS DO ANEXO"])

        logging.info("Script iniciado.")

        log_data = {"Subject": [], "Status": []}

        try:
            extrair_emails_thread(account, selected_folder, progress_var, log_data)
        except Exception as e:
            logging.error(f"Erro durante a execução do script: {str(e)}")

        log_df = pd.DataFrame(log_data)
        log_sheet.append(["Subject", "Status"])
        for index, row in log_df.iterrows():
            log_sheet.append([row["Subject"], row["Status"]])

        # Atualizar a barra de progresso para 100%
        progress_var.update(progress_max)

        criar_planilha_e_log(data_mail_folder, nome_arquivo, main_sheet, log_sheet)

if __name__ == "__main__":
    main()